# Helper Script #1: Shared Functions

# This script contains a collection of functions that are used by multiple scripts
# (The main purpose of these scripts is to read files and extract paths from columns 
#  in the general filepath spreadsheets)



#### SETUP ####

import subprocess
import re
import pandas as pd
import openpyxl
import os
import geopandas as gpd
#from arcgis.gis import GIS
from arcgis.features import FeatureLayer, FeatureLayerCollection, GeoAccessor, GeoSeriesAccessor
#import arcgis.geometry as geom
import numpy as np
import requests



#### FUNCTIONS ####

def wsReadFile (ws: pd.DataFrame, fieldNames: list) -> pd.DataFrame:
    """
    This is a wrapper function for readFile() that initiates the
    process for reading in an XLSX, CSV, TSV, or regular type of file

    'ws' contains path information for the selected watershed

    'fieldNames' is a list of strings that identifies the fields in 'ws' 
    for which file information must be extracted 
    """


    # To start, use 'ws' and extract the file parameters related to the items in 'fieldNames' 
    FILEPATH_PARAMETERS = extractPathInfo(ws, fieldNames)


    # The next step is to determine what kind of file will be read
    # Perform that operation in a separate file
    fileType = guessFileType(FILEPATH_PARAMETERS)  


    # Finally, call the general function readFile() to read in the file
    return readFile(FILEPATH_PARAMETERS, fileType)



def guessFileType (FILEPATH_PARAMETERS: list) -> str:
    """
    Given a list that contains filepath information, try to determine the identity of the file.
    It will be labeled as "XLSX", "CSV", "TSV", "HTML", or "TXT"
    """


    # Start by checking the file extension at the end of the filepath
    # (The filepath should be the first element of the list)
    if re.search("\\.csv$", FILEPATH_PARAMETERS[0], flags = re.IGNORECASE) is not None:
        fileType = "CSV"

    elif re.search("\\.xlsx?$", FILEPATH_PARAMETERS[0], flags = re.IGNORECASE) is not None:
        fileType = "XLSX"

    elif re.search("\\.tsv$", FILEPATH_PARAMETERS[0], flags = re.IGNORECASE) is not None:
        fileType = "TSV"

    elif isURL(FILEPATH_PARAMETERS[0]):
        fileType = "HTML"

    # In all other cases, treat the file as a text file
    # (As a result, the file will be read line-by-line)
    else:
        fileType = "TXT"


    # Return 'fileType'
    return fileType



def wsReadGIS (ws: pd.DataFrame, fieldNames: list) -> gpd.GeoDataFrame:
    """
    This is a precursor function for readFile() that initiates the 
    process for reading in a GIS file

    'ws' contains path information for the selected watershed

    'fieldNames' is a list of strings that identifies the fields in 'ws' 
    for which file information must be extracted 
    """


    # Use 'ws' to extract the file parameters related to the items in 'fieldNames' 
    FILEPATH_PARAMETERS = extractPathInfo(ws, fieldNames)


    # Then, call the general function readFile() to read in the "GIS" file
    return readFile(FILEPATH_PARAMETERS, "GIS")



def readFile (FILEPATH_PARAMETERS: list, FILE_TYPE: str) -> pd.DataFrame | gpd.GeoDataFrame:
    """
    This function attempts to read a file ('CSV', 'XLSX', 'TSV', 'HTML', or GIS files)

    'FILEPATH_PARAMETERS' is a list that contains the relevant filepath components 
    (e.g., filepath and worksheet name for spreadsheets)

    'FILE_TYPE' identifies the type of file and should be 'CSV', 'XLSX', 'TSV', 'HTML', or 'GIS'
    """


    # First, verify that 'FILE_TYPE' contains a valid value
    if FILE_TYPE not in ["GIS", "XLSX", "CSV", "TSV", "HTML", "TXT"]:
        raise Exception("This function can only read 'XLSX', 'CSV', 'TSV', 'HTML', 'TXT', or GIS files. " +
                        "The 'FILE_TYPE' parameter must be one of these values.")
    

    # If there are no issues, identify the file location next
    # The filepath should be a SharePoint path, local path, or URL
    fileLocation = classifyPath(FILEPATH_PARAMETERS[0])


    # Verify the result
    if fileLocation not in ["SharePoint", "Local", "URL"]:
        raise Exception("There is a problem with the function classifyPath()." + 
                        "It should return one of three values ('SharePoint', " +
                        "'Local', or 'URL').")
    
    
    # Then, apply different functions depending on the type of file
    if FILE_TYPE == "XLSX":
        return readXLSX(FILEPATH_PARAMETERS, fileLocation == "SharePoint")

    elif FILE_TYPE == "CSV":
        return readCSV(FILEPATH_PARAMETERS, fileLocation == "SharePoint", sep = ',')
    
    elif FILE_TYPE == "TSV":
        return readCSV(FILEPATH_PARAMETERS, fileLocation == "SharePoint", sep = '\t')

    elif FILE_TYPE == "HTML":
        return readHTML(FILEPATH_PARAMETERS)
    
    elif FILE_TYPE == "TXT":
        return readLines(FILEPATH_PARAMETERS, fileLocation == "SharePoint")

    else: 
        return readGIS(FILEPATH_PARAMETERS, 
                      fileLocation == "SharePoint", fileLocation == "URL")



def classifyPath (filePath: str) -> str:
    """
    Given a filepath, interpret the filepath as a SharePoint path, local path, or URL
    """


    # If the path starts with "http://" or "https://", assume it is a URL
    if isURL(filePath):
        return "URL"

    # Then, try to interpret the filepath as a SharePoint path
    # If a valid match exists, then it is a SharePoint path
    elif os.path.exists(makeSharePointPath(filePath)):
        return "SharePoint"
    
    # After that, try to interpret the filepath as a local path
    # It can be either a relative path or an absolute path
    # Either way, ensure that it exists before assigning this label
    elif os.path.exists(filePath):
        return "Local"

    # Throw an error in all other cases
    else:
        raise ValueError("The file path '" + filePath + "' does not appear to exist")



def isURL (filePath: str) -> bool:
    """
    This function checks if a filepath string is a URL

    If "http://" or "https://" appears at the start of the string,
    it is assumed to be a URL

    Note: re.match only does pattern matching at the start of the string (unlike re.search)
    """

    return re.match("https?://", filePath) is not None


def makeSharePointPath (filePathFragment: str) -> str:
    """
    Given 'filePathFragment' (most of the filepath), write a complete filepath to the file
    
    'filePathFragment' should continue from the SharePoint directory onwards 
    Everything up to "Supply and Demand Assessment - Documents" (inclusive) will already be specified by this function
    
    (This function assumes that the SharePoint filepath is "C:/Users/[username]/Water Boards/Supply and Demand Assessment - Documents/...")
    (In that case, the only unknown variable is the username. Command Prompt will be used to get that information.)
    """


    # Run 'whoami' via Command Prompt and capture the output
    username = subprocess.run("whoami", text = True, stdout = subprocess.PIPE)
    
    
    # Extract the string from the "stdout" element and remove the newline
    username = username.stdout.strip()
    
    
    # The username will be in the format "epa\username"
    # Remove the "epa\" portion
    username = re.sub("^epa\\\\", "", username)

    
    # Return a full string using 'username' and 'filePathFragment'
    # (Essentially "C:/Users/" + username + "/Water Boards/Supply and Demand Assessment - Documents/" + filePathFragment)
    return os.path.join("C:/Users", username, "Water Boards/Supply and Demand Assessment - Documents", filePathFragment)



def readXLSX (FILEPATH_PARAMETERS: list, isSharePoint: bool) -> pd.DataFrame:
    """
    Read in an Excel file as a data frame

    'FILEPATH_PARAMETERS' is a list that contains the path to a spreadsheet
    (and optionally, a worksheet name)

    'isSharePoint' is a boolean that indicates whether the filepath is a SharePoint path
    """


    # If 'isSharePoint' is True, convert the first element of 'FILEPATH_PARAMETERS' into a SharePoint path
    if isSharePoint == True:
        filePath = makeSharePointPath(FILEPATH_PARAMETERS[0])
    
    # If not, use it as is
    else:
        filePath = FILEPATH_PARAMETERS[0]


    # If a second element is present in 'FILEPATH_PARAMETERS',
    # it is probably a string containing the worksheet name
    if len(FILEPATH_PARAMETERS) > 1 and FILEPATH_PARAMETERS[1] is not None:
        worksheetName = FILEPATH_PARAMETERS[1]

    # If no second element is present, by default, read in the first worksheet of the file
    else:
        worksheetName = 0


    # After that, use read_excel to read in the Excel table
    # (Watch for errors)
    try:
        xlsxDF = pd.read_excel(filePath, sheet_name = worksheetName)

    # Output a custom message for one of the most commonly encountered issues
    # (Reading in a spreadsheet when the file is open)
    except PermissionError as err:
        raise PermissionError("The most common cause of this error is having the spreadsheet open when running the scripts. " + 
                  "Please make sure '" + filePath + "' is closed!")


    # If there are no issues, return 'xlsxDF'
    return xlsxDF



def readCSV (FILEPATH_PARAMETERS: list, isSharePoint: bool, sep: str) -> pd.DataFrame:
    """
    Read in a CSV (or similar) file as a data frame

    'FILEPATH_PARAMETERS' is a list that contains the path to the file
    (there should be no other parameters)

    'isSharePoint' is a boolean that indicates whether the filepath is a SharePoint path
    """

      # If 'isSharePoint' is True, convert the first element of 'FILEPATH_PARAMETERS' into a SharePoint path
    if isSharePoint == True:
        filePath = makeSharePointPath(FILEPATH_PARAMETERS[0])
    
    # If not, use it as is
    else:
        filePath = FILEPATH_PARAMETERS[0]


    # After that, try to read in the data using read_csv
    csvDF = pd.read_csv(filePath, sep = sep)


    # If there are no issues, return 'csvDF'
    return csvDF



def readHTML (FILEPATH_PARAMETERS: list) -> pd.DataFrame:
    """
    Read in an HTML file as a data frame using an HTTP GET request

    'FILEPATH_PARAMETERS' is a list that contains the path to the file 
    (there should only be one element)
    """


    # Try to read in the data using a GET request
    htmlRes = requests.get(FILEPATH_PARAMETERS[0])


    # Verify that the request was successful
    if htmlRes.status_code != 200:
        raise ValueError("The URL '" + FILEPATH_PARAMETERS[0] + "' did not return a valid response to the script's GET request")


    # Format 'htmlRes' as a data frame (set the column name to "FILE_CONTENTS")
    htmlDF = pd.DataFrame(htmlRes, columns = ["FILE_CONTENTS"])


    # If there are no issues, return 'csvDF'
    return htmlDF



def readLines (FILEPATH_PARAMETERS: list, isSharePoint: bool) -> pd.DataFrame:
    """
    Read in a text file as a data frame

    'FILEPATH_PARAMETERS' is a list that contains the path to the file
    (there should be no other parameters)

    'isSharePoint' is a boolean that indicates whether the filepath is a SharePoint path
    """

      # If 'isSharePoint' is True, convert the first element of 'FILEPATH_PARAMETERS' into a SharePoint path
    if isSharePoint == True:
        filePath = makeSharePointPath(FILEPATH_PARAMETERS[0])
    
    # If not, use it as is
    else:
        filePath = FILEPATH_PARAMETERS[0]


    # After that, try to read in the data
    with open(filePath, 'r', encoding = "utf-8") as f:
        txtDF = f.readlines()


    # 'txtDF' needs to be formatted as a data frame next
    # Save it as a single-column data frame with the default column name "FILE_CONTENTS"
    txtDF = pd.DataFrame(txtDF, columns = ["FILE_CONTENTS"])


    # If there are no issues, return 'txtDF'
    return txtDF



def readGIS (FILEPATH_PARAMETERS: list, isSharePoint: bool, isURL: bool) -> gpd.GeoDataFrame:
    """
    Read in a GIS file as a data frame

    It can be something like a shapefile or a geoJSON file
    # Alternatively, it can be a layer in a container (like a geopackage or geodatabase)

    A URL that points to an ArcGIS Portal layer can also be provided

    'FILEPATH_PARAMETERS' is a list that contains the path to the file
    (along with the layer name, if separate from the container path)

    'isSharePoint' is a boolean that indicates whether the filepath is a SharePoint path

    'isURL' is a boolean that indicates whether the path is an ArcGIS Portal link
    """

      # If 'isSharePoint' is True, convert the first element of 'FILEPATH_PARAMETERS' into a SharePoint path
    if isSharePoint == True:
        filePath = makeSharePointPath(FILEPATH_PARAMETERS[0])
    
    # If not, use it as is
    else:
        filePath = FILEPATH_PARAMETERS[0]


    # After that, consider whether a second element is present in 'FILEPATH_PARAMETERS'
    # If yes, it is probably a string containing the layer name
    if len(FILEPATH_PARAMETERS) > 1 and FILEPATH_PARAMETERS[1] is not None:
        layerName = FILEPATH_PARAMETERS[1]

    # If no second element is present, by default, set 'layerName' to None
    else:
        layerName = None


    # After that, for ArcGIS Portal URLs, try to read in the layer using their functions
    # (And then convert it into a GeoDataFrame)
    if isURL and isPortalURL(filePath, layerName):
        gisDF = getPortalLayer(filePath, layerName)

    else:
        # For all other cases, try to use read_file
        gisDF = gpd.read_file(filePath, layer = layerName)


    # If there are no issues, return 'gisDF'
    return gisDF



def isPortalURL(urlPath: str, layerName: str | None) -> bool:
    """
    This function tries to check if a URL is a valid Portal link
    It will return either True or False

    'urlPath' contains the URL that points to a GIS layer
    (It should either be a path to a "Feature Layer" or a "Feature Layer Collection")

    'layerName' is the Feature Layer index (if 'urlPath' points to a Feature Layer Collection)
    """


    # Check for keywords like "ArcGIS" and "Portal" in the URL
    # If neither appears in 'urlPath', return False
    if re.search("ArcGIS|Portal", urlPath, flags = re.IGNORECASE) is None:
        return False

    
    # If 'layerName' is not empty, verify that it is a numeric value (just digits)
    # (For a Feature Layer Collection, 'layerName' must be the index that corresponds to a layer)
    if layerName is not None and re.search("^[0-9]+$", layerName) is None:
        return False


    # In all other scenarios, return True
    return True



def getPortalLayer(urlPath: str, layerName: str | None) -> gpd.GeoDataFrame:
    """
    This function reads a Feature Layer or Feature Layer Collection from ArcGIS,
    extracts a layer, and tries to read the layer as a GeoDataFrame
    """
    
    
    # If 'layerName' is NOT empty, assume 'urlPath' points to a Feature Layer Collection
    # Otherwise, assume that 'urlPath' points to a Feature Layer

    # Either way, extract a layer from the URL as 'layerDF'
    if layerName is None:
        layerDF = FeatureLayer(urlPath)

    else:
        layerCollection = FeatureLayerCollection(urlPath)

        layerDF = layerCollection.layers[int(layerName)]


    # This next procedure is a hacky method to get the ArcGIS layer into a GeoPandas GeoDataFrame
    # (Inspired by https://gis.stackexchange.com/questions/418040/converting-arcgis-spatially-enabled-dataframe-to-geojson)

    
    # Convert 'layerDF' into a spatially-enabled data frame
    layerDF = pd.DataFrame.spatial.from_layer(layerDF)


    # After that, convert 'layerDF' into a FeatureSet object
    layerDF = layerDF.spatial.to_featureset()


    # Write 'layerDF' to a temporary geoJSON file
    tempFileName = "tempScript.geojson"


    with open(tempFileName, "w", encoding = "utf-8") as file:
        file.write(layerDF.to_geojson)

    
    # Read in this temporary file using read_file
    # (It is now a GeoDataFrame)
    gisDF = gpd.read_file(tempFileName)


    # Delete the temporary file
    os.unlink(tempFileName)


    # Finally, return 'gisDF'
    return gisDF



def extractPathInfo(ws: pd.DataFrame, fieldNames: list) -> list:
    """
    This function extracts information from 'ws' for the data fields specified in 'fieldNames' 
    """


    # 'ws' has one of two formats:

    #   (1) With a column-based formatting, each watershed has its own column and
    #       each row in 'ws' corresponds to a different field parameter
    
    #   (2) With a row-based formatting, each watershed has its own row and
    #       each column in 'ws' corresponds to a different field parameter
    
    # Check ws.shape() to distinguish between the two formats
    # (This distinction will affect how the function extracts path information from 'ws')
    if ws.shape[0] == 1:
        spreadsheetStyle = "ROW"
    else:
        spreadsheetStyle = "COLUMN"


    # The next step is to actually extract the values for each of the names given in 'fieldNames'


    # Define a list to hold this information
    parameterList = []


    # For every parameter name that appears in 'fieldNames',
    # add the watershed's corresponding value to 'parameterList' 
    for field in fieldNames:

        # Most of the loop's actions will different depending on 'spreadsheetStyle'
        if spreadsheetStyle == "COLUMN":
            
            # First, make sure 'field' actually exists in 'ws'
            if len(ws[ws.iloc[:, 0] == field]) == 0:
                raise ValueError("The field '" + field + "' does not exist in 'ws'!")


            # If yes, extract the second column's value (index 1) from the row whose first column (index 0) matches 'field'
            extractedValue = ws[ws.iloc[:, 0] == field].iloc[0, 1]

        # For a row-based spreadsheet, the process is the same, but the extraction code is different
        elif spreadsheetStyle == "ROW":

            # As before, confirm that 'field' actually exists in 'ws'
            if sum(ws.index.values == field) == 0:
                raise ValueError("The field '" + field + "' does not exist in 'ws'!")


            # If 'field' is present, extract the row's value from the column whose index name matches 'field'
            extractedValue = ws[ws.index.values == field].iloc[0]

        # If 'spreadsheetStyle' is neither "ROW" nor "COLUMN", somebody made a mistake in this function
        else:
            raise ValueError('The variable "spreadsheetStyle" was not set correctly in the function extractPathInfo()!')


        # At this point in the loop through 'fieldNames', whether 'spreadsheetStyle' is "ROW" or "COLUMN",
        # 'extractedValue' has been obtained (as a string, most likely)
        # An important check now is whether 'extractedValue' is NaN
        # (This would happen if its entry in the paths spreadsheet is empty)

        # If 'extractedValue' is NOT NaN, add that element to 'parameterList'
        if not np.isnan(extractedValue):
            parameterList.append(extractedValue)
    

    # After the loop is complete, if 'parameterList' is still empty, raise an exception
    if len(parameterList) == 0:
        raise ValueError("The watershed lacks path information for all of the following field(s): " + "; ".join(fieldNames))

    
    # Finally, return 'parameterList'
    return parameterList
